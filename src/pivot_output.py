"""
Generate a pivot-style schedule: rows = employees, columns = days of the week.
Each cell shows the station assignment (AM & PM combined).
"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from kitchen_data import DAYS, SERVICE_LEVELS, EMPLOYEES, OPEN_DAYS
from scheduler import run_scheduler
from models import ServiceLevel

HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
DAY_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
OFF_FILL = PatternFill("solid", fgColor="E8E8E8")
CLOSED_FILL = PatternFill("solid", fgColor="D9D9D9")
AM_FILL = PatternFill("solid", fgColor="DAEEF3")
PM_FILL = PatternFill("solid", fgColor="FDE9D9")
LEARN_FILL = PatternFill("solid", fgColor="FFF2CC")
TRAINING_FILL = PatternFill("solid", fgColor="E2CFEA")  # light purple — training shadow
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def generate_pivot(output_path):
    assignments, shift_counts = run_scheduler()

    # Build lookup: (employee, day) -> list of (shift, label, coverage_type)
    lookup = {}
    for a in assignments:
        key = (a.employee, a.day)
        label = f"{a.station}"
        if a.coverage_type == "training":
            label += " 🎓 (shadow training)"
        elif a.coverage_type == "learning":
            label += " ⚠️"
        # If employee also has a training_on note (e.g. Sam: Pasta + training Sauté),
        # append it so the cell makes the dual role visible.
        if a.notes and "Also training on" in a.notes:
            training_stations = a.notes.replace("Also training on ", "")
            label += f"\n▶ training: {training_stations}"
        lookup.setdefault(key, []).append((a.shift, label, a.coverage_type))

    # Determine employee display order: leadership, pm_staff, am_staff
    role_order = {"leadership": 0, "pm_staff": 1, "am_staff": 2}
    sorted_employees = sorted(EMPLOYEES, key=lambda e: (role_order.get(e.role, 9), e.name))

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"
    ws.sheet_properties.tabColor = "2F5496"

    # ── Row 1: Title ────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = "Acquerello Weekly Schedule — Week of 2026-03-02"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: Service levels ───────────────────────────────────
    ws.cell(row=2, column=1, value="Service Level").font = Font(name="Arial", bold=True, size=9, color="808080")
    ws.cell(row=2, column=1).alignment = CENTER
    for ci, day in enumerate(DAYS, 2):
        sl = SERVICE_LEVELS[day]
        cell = ws.cell(row=2, column=ci, value=sl.value.upper())
        cell.font = Font(name="Arial", bold=True, size=9,
                         color="C00000" if sl == ServiceLevel.PEAK else
                               "808080" if sl == ServiceLevel.CLOSED else "BF8F00")
        cell.alignment = CENTER

    # ── Row 3: Header (Employee | Mon | Tue | ... | Sun) ────────
    header_row = 3
    ws.cell(row=header_row, column=1, value="Employee")
    for ci, day in enumerate(DAYS, 2):
        ws.cell(row=header_row, column=ci, value=day[:3])
    for c in range(1, len(DAYS) + 2):
        cell = ws.cell(row=header_row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # ── Data rows ───────────────────────────────────────────────
    current_role = None
    row = header_row + 1
    for emp in sorted_employees:
        # Role separator
        if emp.role != current_role:
            current_role = emp.role
            role_label = {"leadership": "LEADERSHIP", "pm_staff": "PM LINE", "am_staff": "AM PREP"}.get(current_role, current_role)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DAYS) + 1)
            cell = ws.cell(row=row, column=1, value=role_label)
            cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="808080")
            cell.alignment = CENTER
            for c in range(1, len(DAYS) + 2):
                ws.cell(row=row, column=c).border = BORDER
            row += 1

        # Employee name
        ws.cell(row=row, column=1, value=emp.name).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=1).border = BORDER

        for ci, day in enumerate(DAYS, 2):
            cell = ws.cell(row=row, column=ci)
            cell.alignment = CENTER
            cell.border = BORDER
            cell.font = BODY_FONT

            sl = SERVICE_LEVELS[day]
            if sl == ServiceLevel.CLOSED:
                cell.value = "CLOSED"
                cell.fill = CLOSED_FILL
                cell.font = Font(name="Arial", size=9, color="999999")
                continue

            entries = lookup.get((emp.name, day), [])
            if not entries:
                cell.value = "OFF"
                cell.fill = OFF_FILL
                cell.font = Font(name="Arial", size=10, color="999999", italic=True)
            else:
                parts = []
                has_learn = False
                has_training = False
                has_am = False
                has_pm = False
                for shift, label, cov in entries:
                    parts.append(label)
                    if cov == "training":
                        has_training = True
                    elif cov == "learning":
                        has_learn = True
                    if shift == "AM":
                        has_am = True
                    if shift == "PM":
                        has_pm = True
                cell.value = "\n".join(parts)
                if has_training:
                    cell.fill = TRAINING_FILL
                elif has_learn:
                    cell.fill = LEARN_FILL
                elif has_am and not has_pm:
                    cell.fill = AM_FILL
                elif has_pm and not has_am:
                    cell.fill = PM_FILL
                else:
                    cell.fill = PatternFill("solid", fgColor="EBF1DE")  # both

        row += 1

    # ── Legend row ───────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="Legend:").font = Font(name="Arial", bold=True, size=9)
    legends = [
        (2, "AM shift", AM_FILL),
        (3, "PM shift", PM_FILL),
        (4, "Learning ⚠️", LEARN_FILL),
        (5, "Training 🎓", TRAINING_FILL),
        (6, "OFF", OFF_FILL),
        (7, "CLOSED", CLOSED_FILL),
    ]
    for col, text, fill in legends:
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill
        c.font = Font(name="Arial", size=9)
        c.alignment = CENTER
        c.border = BORDER

    # ── Column widths ───────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    for ci in range(2, len(DAYS) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    # Row heights for wrapped text
    for r in range(header_row + 1, row):
        ws.row_dimensions[r].height = 32

    wb.save(output_path)
    print(f"Pivot schedule → {output_path}")


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.dirname(__file__)), "schedule_output.xlsx")
    if len(sys.argv) > 1:
        out = sys.argv[1]
    generate_pivot(out)
