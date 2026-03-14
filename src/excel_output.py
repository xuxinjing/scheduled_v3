"""
Generate schedule_output.xlsx with required worksheets:
Instructions, Employees, Station_Coverage, Daily_Schedule, Validation_Input, Summary
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from kitchen_data import OPEN_DAYS, SERVICE_LEVELS, EMPLOYEES, AM_STATIONS, PM_STATIONS
from scheduler import run_scheduler, build_pm_requirements, build_am_requirements
from validator import validate


# ── Style constants ─────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
OFF_FILL = PatternFill("solid", fgColor="F2F2F2")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4EC")
PASS_FILL = PatternFill("solid", fgColor="E8F5E9")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _style_body(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = LEFT_WRAP if c > 1 else CENTER
            cell.border = BORDER


def _auto_width(ws, ncols, min_w=12, max_w=28):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min_w


def generate(output_path: str):
    assignments, shift_counts = run_scheduler()
    report = validate(assignments, shift_counts)

    wb = Workbook()

    # ═══ 1. Instructions ════════════════════════════════════════
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "2F5496"
    lines = [
        ["Acquerello Kitchen Schedule — Week of 2026-03-09"],
        [""],
        ["This workbook contains the weekly schedule generated from kitchen_state.md and week_constraints.md."],
        [""],
        ["Sheets:"],
        ["  Employees — staff roster with capabilities"],
        ["  Station_Coverage — required stations per day/service level"],
        ["  Daily_Schedule — the actual schedule assignments"],
        ["  Validation_Input — constraint data used by the validator"],
        ["  Summary — shift counts, coverage overview, and warnings"],
        [""],
        [f"Validation Status: {report['status']}"],
        [f"Errors: {len(report['errors'])}  |  Warnings: {len(report['warnings'])}"],
        [""],
        ["Key:"],
        ["  stable = fully independent coverage"],
        ["  learning = training-level, needs oversight"],
        ["  emergency = last-resort coverage only"],
    ]
    for i, row in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=row[0]).font = Font(name="Arial", size=11,
            bold=(i == 1))
    ws.column_dimensions["A"].width = 80

    # ═══ 2. Employees ══════════════════════════════════════════
    ws2 = wb.create_sheet("Employees")
    ws2.sheet_properties.tabColor = "548235"
    headers = ["Name", "Role", "Capabilities", "Preferred Stations", "Unavailable"]
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))
    for e in EMPLOYEES:
        caps = ", ".join(f"{s} ({c.value})" for s, c in e.capabilities.items())
        prefs = ", ".join(e.preferred_stations) if e.preferred_stations else "—"
        unavail = ", ".join(e.unavailable_days) if e.unavailable_days else "—"
        ws2.append([e.name, e.role, caps, prefs, unavail])
    _style_body(ws2, 2, ws2.max_row, len(headers))
    for c, w in enumerate([14, 12, 50, 24, 20], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ═══ 3. Station_Coverage ═══════════════════════════════════
    ws3 = wb.create_sheet("Station_Coverage")
    ws3.sheet_properties.tabColor = "BF8F00"
    headers = ["Day", "Service Level", "Shift", "Station", "Required HC"]
    ws3.append(headers)
    _style_header(ws3, 1, len(headers))
    for day in OPEN_DAYS:
        sl = SERVICE_LEVELS[day].value
        for st, hc in build_am_requirements(day):
            ws3.append([day, sl, "AM", st, hc])
        for st, hc in build_pm_requirements(day):
            ws3.append([day, sl, "PM", st, hc])
    _style_body(ws3, 2, ws3.max_row, len(headers))
    _auto_width(ws3, len(headers))

    # ═══ 4. Daily_Schedule ═════════════════════════════════════
    ws4 = wb.create_sheet("Daily_Schedule")
    ws4.sheet_properties.tabColor = "C00000"
    headers = ["Day", "Service", "Shift", "Station", "Employee", "Coverage Type", "Notes", "Status"]
    ws4.append(headers)
    _style_header(ws4, 1, len(headers))
    for a in assignments:
        sl = SERVICE_LEVELS[a.day].value
        status = "OK"
        if a.coverage_type == "learning":
            status = "WATCH"
        elif a.coverage_type == "emergency":
            status = "RISK"
        ws4.append([a.day, sl, a.shift, a.station, a.employee, a.coverage_type, a.notes, status])
    # Style
    for r in range(2, ws4.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws4.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = LEFT_WRAP
            cell.border = BORDER
        status_cell = ws4.cell(row=r, column=8)
        if status_cell.value == "WATCH":
            status_cell.fill = WARN_FILL
        elif status_cell.value == "RISK":
            status_cell.fill = ERROR_FILL
        else:
            status_cell.fill = PASS_FILL
    for c, w in enumerate([12, 10, 8, 18, 14, 14, 32, 10], 1):
        ws4.column_dimensions[get_column_letter(c)].width = w

    # ═══ 5. Validation_Input ═══════════════════════════════════
    ws5 = wb.create_sheet("Validation_Input")
    ws5.sheet_properties.tabColor = "7030A0"
    headers = ["Category", "Detail"]
    ws5.append(headers)
    _style_header(ws5, 1, 2)
    for e in report["errors"]:
        ws5.append(["ERROR", e])
    for w in report["warnings"]:
        ws5.append(["WARNING", w])
    for a in report["assumptions"]:
        ws5.append(["ASSUMPTION", a])
    _style_body(ws5, 2, ws5.max_row, 2)
    for r in range(2, ws5.max_row + 1):
        cat = ws5.cell(row=r, column=1).value
        if cat == "ERROR":
            ws5.cell(row=r, column=1).fill = ERROR_FILL
        elif cat == "WARNING":
            ws5.cell(row=r, column=1).fill = WARN_FILL
    ws5.column_dimensions["A"].width = 14
    ws5.column_dimensions["B"].width = 70

    # ═══ 6. Summary ════════════════════════════════════════════
    ws6 = wb.create_sheet("Summary")
    ws6.sheet_properties.tabColor = "00B050"
    ws6.append(["Acquerello Schedule Summary — Week of 2026-03-09"])
    ws6.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=13)
    ws6.merge_cells("A1:D1")

    # Shift counts
    ws6.append([])
    ws6.append(["Employee", "Total Shifts", "Role"])
    _style_header(ws6, 3, 3)
    sorted_counts = sorted(shift_counts.items(), key=lambda x: -x[1])
    for name, cnt in sorted_counts:
        emp = next((e for e in EMPLOYEES if e.name == name), None)
        role = emp.role if emp else "?"
        ws6.append([name, cnt, role])
    _style_body(ws6, 4, ws6.max_row, 3)

    # Day-off summary for PM staff
    r = ws6.max_row + 2
    ws6.cell(row=r, column=1, value="PM Staff Days Off").font = SUBHEADER_FONT
    r += 1
    ws6.append(["Employee", "Days Working", "Days Off"])
    _style_header(ws6, r, 3)
    pm_names = {e.name for e in EMPLOYEES if e.role in ("leadership", "pm_staff")}
    by_emp = {}
    for a in assignments:
        if a.employee in pm_names and a.shift == "PM":
            by_emp.setdefault(a.employee, set()).add(a.day)
    for name in sorted(pm_names):
        working = by_emp.get(name, set())
        off = set(OPEN_DAYS) - working
        ws6.append([name, ", ".join(sorted(working, key=OPEN_DAYS.index)),
                     ", ".join(sorted(off, key=OPEN_DAYS.index)) if off else "—"])
    _style_body(ws6, r + 1, ws6.max_row, 3)

    # Validation summary
    r = ws6.max_row + 2
    ws6.cell(row=r, column=1, value=f"Validation: {report['status']}").font = Font(
        name="Arial", bold=True, size=12,
        color="C00000" if report["errors"] else "00B050")
    r += 1
    ws6.cell(row=r, column=1, value=f"Errors: {len(report['errors'])}").font = BODY_FONT
    r += 1
    ws6.cell(row=r, column=1, value=f"Warnings: {len(report['warnings'])}").font = BODY_FONT

    for c, w in enumerate([18, 36, 36, 20], 1):
        ws6.column_dimensions[get_column_letter(c)].width = w

    wb.save(output_path)
    return report


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "schedule_output.xlsx"
    report = generate(out)
    print(f"Schedule generated → {out}")
    print(f"Status: {report['status']}")
    print(f"Errors: {len(report['errors'])}, Warnings: {len(report['warnings'])}")
    for e in report["errors"]:
        print(f"  {e}")
    for w in report["warnings"]:
        print(f"  {w}")
