"""Pivot preview and workbook generation."""
from __future__ import annotations

from io import BytesIO

from .context import KitchenContext
from .models import ServiceLevel


def build_pivot_preview(context: KitchenContext, assignments) -> dict:
    lookup = {}
    for assignment in assignments:
        key = (assignment.employee, assignment.day)
        label = assignment.station
        if assignment.coverage_type == "training":
            label += " (shadow training)"
        elif assignment.coverage_type == "learning":
            label += " (!)"
        if assignment.notes and "Also training on" in assignment.notes:
            label += f" | {assignment.notes}"
        lookup.setdefault(key, []).append(
            {
                "shift": assignment.shift,
                "label": label,
                "coverage_type": assignment.coverage_type,
            }
        )

    role_order = {"leadership": 0, "pm_staff": 1, "am_staff": 2}
    rows = []
    for employee in sorted(context.employees, key=lambda item: (role_order.get(item.role, 9), item.name)):
        cells = {}
        for day in context.days:
            service_level = context.service_levels[day]
            if service_level == ServiceLevel.CLOSED:
                cells[day] = {"text": "CLOSED", "entries": []}
                continue
            entries = lookup.get((employee.name, day), [])
            cells[day] = {
                "text": " / ".join(entry["label"] for entry in entries) if entries else "OFF",
                "entries": entries,
            }
        rows.append({"employee": employee.name, "role": employee.role, "cells": cells})
    return {"week_start": context.week_start, "days": context.days, "rows": rows}


def generate_pivot_workbook_bytes(context: KitchenContext, assignments) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to generate pivot workbooks") from exc

    preview = build_pivot_preview(context, assignments)

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    body_font = Font(name="Arial", size=10)
    off_fill = PatternFill("solid", fgColor="E8E8E8")
    closed_fill = PatternFill("solid", fgColor="D9D9D9")
    am_fill = PatternFill("solid", fgColor="DAEEF3")
    pm_fill = PatternFill("solid", fgColor="FDE9D9")
    learn_fill = PatternFill("solid", fgColor="FFF2CC")
    training_fill = PatternFill("solid", fgColor="E2CFEA")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Weekly Schedule"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = f"Acquerello Weekly Schedule - Week of {context.week_start}"
    sheet["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    sheet["A1"].alignment = center

    sheet.cell(row=2, column=1, value="Service Level").font = Font(name="Arial", bold=True, size=9, color="808080")
    sheet.cell(row=2, column=1).alignment = center
    for column_index, day in enumerate(context.days, 2):
        service_level = context.service_levels[day]
        cell = sheet.cell(row=2, column=column_index, value=service_level.value.upper())
        cell.font = Font(
            name="Arial",
            bold=True,
            size=9,
            color="C00000" if service_level == ServiceLevel.PEAK else "808080" if service_level == ServiceLevel.CLOSED else "BF8F00",
        )
        cell.alignment = center

    header_row = 3
    sheet.cell(row=header_row, column=1, value="Employee")
    for column_index, day in enumerate(context.days, 2):
        sheet.cell(row=header_row, column=column_index, value=day[:3])
    for column_index in range(1, len(context.days) + 2):
        cell = sheet.cell(row=header_row, column=column_index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    current_role = None
    row_index = header_row + 1
    for row in preview["rows"]:
        if row["role"] != current_role:
            current_role = row["role"]
            role_label = {"leadership": "LEADERSHIP", "pm_staff": "PM LINE", "am_staff": "AM PREP"}.get(current_role, current_role)
            sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(context.days) + 1)
            cell = sheet.cell(row=row_index, column=1, value=role_label)
            cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="808080")
            cell.alignment = center
            for column_index in range(1, len(context.days) + 2):
                sheet.cell(row=row_index, column=column_index).border = border
            row_index += 1

        sheet.cell(row=row_index, column=1, value=row["employee"]).font = Font(name="Arial", bold=True, size=10)
        sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="left", vertical="center")
        sheet.cell(row=row_index, column=1).border = border

        for column_index, day in enumerate(context.days, 2):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = center
            cell.border = border
            cell.font = body_font

            service_level = context.service_levels[day]
            if service_level == ServiceLevel.CLOSED:
                cell.value = "CLOSED"
                cell.fill = closed_fill
                cell.font = Font(name="Arial", size=9, color="999999")
                continue

            cell_data = row["cells"][day]
            cell.value = cell_data["text"]
            entries = cell_data["entries"]
            if not entries:
                cell.fill = off_fill
                cell.font = Font(name="Arial", size=10, color="999999", italic=True)
                continue

            coverage_types = {entry["coverage_type"] for entry in entries}
            shifts = {entry["shift"] for entry in entries}
            if "training" in coverage_types:
                cell.fill = training_fill
            elif "learning" in coverage_types:
                cell.fill = learn_fill
            elif shifts == {"AM"}:
                cell.fill = am_fill
            elif shifts == {"PM"}:
                cell.fill = pm_fill
            else:
                cell.fill = PatternFill("solid", fgColor="EBF1DE")
        row_index += 1

    legend_row = row_index + 1
    sheet.cell(row=legend_row, column=1, value="Legend:").font = Font(name="Arial", bold=True, size=9)
    legends = [
        (2, "AM shift", am_fill),
        (3, "PM shift", pm_fill),
        (4, "Learning (!)", learn_fill),
        (5, "Training", training_fill),
        (6, "OFF", off_fill),
        (7, "CLOSED", closed_fill),
    ]
    for column_index, text, fill in legends:
        cell = sheet.cell(row=legend_row, column=column_index, value=text)
        cell.fill = fill
        cell.font = Font(name="Arial", size=9)
        cell.alignment = center
        cell.border = border

    sheet.column_dimensions["A"].width = 14
    for column_index in range(2, len(context.days) + 2):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18

    for row_number in range(header_row + 1, legend_row):
        sheet.row_dimensions[row_number].height = 32

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
