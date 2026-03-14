"""Workbook generation that returns bytes instead of writing files."""
from __future__ import annotations

from io import BytesIO

from .context import KitchenContext


def generate_schedule_workbook_bytes(context: KitchenContext, assignments, shift_counts: dict[str, int], report: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to generate Excel artifacts") from exc

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    subheader_font = Font(name="Arial", bold=True, size=10)
    body_font = Font(name="Arial", size=10)
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="FCE4EC")
    pass_fill = PatternFill("solid", fgColor="E8F5E9")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header(worksheet, row: int, columns: int) -> None:
        for column_index in range(1, columns + 1):
            cell = worksheet.cell(row=row, column=column_index)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def style_body(worksheet, start_row: int, end_row: int, columns: int) -> None:
        for row_number in range(start_row, end_row + 1):
            for column_index in range(1, columns + 1):
                cell = worksheet.cell(row=row_number, column=column_index)
                cell.font = body_font
                cell.alignment = left_wrap if column_index > 1 else center
                cell.border = border

    workbook = Workbook()

    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.sheet_properties.tabColor = "2F5496"
    lines = [
        f"Acquerello Kitchen Schedule - Week of {context.week_start}",
        "",
        "This workbook contains the weekly schedule generated from restaurant config and week config.",
        "",
        "Sheets:",
        "  Employees - staff roster with capabilities",
        "  Station_Coverage - required stations per day/service level",
        "  Daily_Schedule - the actual schedule assignments",
        "  Validation_Input - constraint data used by the validator",
        "  Summary - shift counts, coverage overview, and warnings",
        "",
        f"Validation Status: {report['status']}",
        f"Errors: {len(report['errors'])}  |  Warnings: {len(report['warnings'])}",
        "",
        "Key:",
        "  stable = fully independent coverage",
        "  learning = training-level, needs oversight",
        "  emergency = last-resort coverage only",
    ]
    for index, text in enumerate(lines, 1):
        instructions.cell(row=index, column=1, value=text).font = Font(name="Arial", size=11, bold=(index == 1))
    instructions.column_dimensions["A"].width = 80

    employees_sheet = workbook.create_sheet("Employees")
    employees_sheet.sheet_properties.tabColor = "548235"
    employee_headers = ["Name", "Role", "Capabilities", "Preferred Stations", "Unavailable", "Forced Days"]
    employees_sheet.append(employee_headers)
    style_header(employees_sheet, 1, len(employee_headers))
    for employee in context.employees:
        capabilities = ", ".join(f"{station} ({capability.value})" for station, capability in employee.capabilities.items()) or "-"
        preferred = ", ".join(employee.preferred_stations) if employee.preferred_stations else "-"
        unavailable = ", ".join(employee.unavailable_days) if employee.unavailable_days else "-"
        forced = ", ".join(employee.forced_days) if employee.forced_days else "-"
        employees_sheet.append([employee.name, employee.role, capabilities, preferred, unavailable, forced])
    style_body(employees_sheet, 2, employees_sheet.max_row, len(employee_headers))
    for column_index, width in enumerate([14, 12, 50, 24, 20, 20], 1):
        employees_sheet.column_dimensions[get_column_letter(column_index)].width = width

    coverage_sheet = workbook.create_sheet("Station_Coverage")
    coverage_sheet.sheet_properties.tabColor = "BF8F00"
    coverage_headers = ["Day", "Service Level", "Shift", "Station", "Required HC"]
    coverage_sheet.append(coverage_headers)
    style_header(coverage_sheet, 1, len(coverage_headers))
    from .scheduler import SchedulerEngine

    engine = SchedulerEngine(context)
    for day in context.open_days:
        service_level = context.service_levels[day].value
        for station, headcount in engine.build_am_requirements(day):
            coverage_sheet.append([day, service_level, "AM", station, headcount])
        for station, headcount in engine.build_pm_requirements(day):
            coverage_sheet.append([day, service_level, "PM", station, headcount])
    style_body(coverage_sheet, 2, coverage_sheet.max_row, len(coverage_headers))

    daily_sheet = workbook.create_sheet("Daily_Schedule")
    daily_sheet.sheet_properties.tabColor = "C00000"
    daily_headers = ["Day", "Service", "Shift", "Station", "Employee", "Coverage Type", "Notes", "Status"]
    daily_sheet.append(daily_headers)
    style_header(daily_sheet, 1, len(daily_headers))
    for assignment in assignments:
        service_level = context.service_levels[assignment.day].value
        status = "OK"
        if assignment.coverage_type == "learning":
            status = "WATCH"
        elif assignment.coverage_type == "emergency":
            status = "RISK"
        daily_sheet.append(
            [
                assignment.day,
                service_level,
                assignment.shift,
                assignment.station,
                assignment.employee,
                assignment.coverage_type,
                assignment.notes,
                status,
            ]
        )
    for row_number in range(2, daily_sheet.max_row + 1):
        for column_index in range(1, len(daily_headers) + 1):
            cell = daily_sheet.cell(row=row_number, column=column_index)
            cell.font = body_font
            cell.alignment = left_wrap
            cell.border = border
        status_cell = daily_sheet.cell(row=row_number, column=8)
        if status_cell.value == "WATCH":
            status_cell.fill = warn_fill
        elif status_cell.value == "RISK":
            status_cell.fill = error_fill
        else:
            status_cell.fill = pass_fill
    for column_index, width in enumerate([12, 10, 8, 18, 14, 14, 32, 10], 1):
        daily_sheet.column_dimensions[get_column_letter(column_index)].width = width

    validation_sheet = workbook.create_sheet("Validation_Input")
    validation_sheet.sheet_properties.tabColor = "7030A0"
    validation_sheet.append(["Category", "Detail"])
    style_header(validation_sheet, 1, 2)
    for error in report["errors"]:
        validation_sheet.append(["ERROR", error])
    for warning in report["warnings"]:
        validation_sheet.append(["WARNING", warning])
    for assumption in report["assumptions"]:
        validation_sheet.append(["ASSUMPTION", assumption])
    style_body(validation_sheet, 2, validation_sheet.max_row, 2)
    for row_number in range(2, validation_sheet.max_row + 1):
        category = validation_sheet.cell(row=row_number, column=1).value
        if category == "ERROR":
            validation_sheet.cell(row=row_number, column=1).fill = error_fill
        elif category == "WARNING":
            validation_sheet.cell(row=row_number, column=1).fill = warn_fill
    validation_sheet.column_dimensions["A"].width = 14
    validation_sheet.column_dimensions["B"].width = 70

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.sheet_properties.tabColor = "00B050"
    summary_sheet.append([f"Acquerello Schedule Summary - Week of {context.week_start}"])
    summary_sheet.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=13)
    summary_sheet.merge_cells("A1:D1")
    summary_sheet.append([])
    summary_sheet.append(["Employee", "Total Shifts", "Role"])
    style_header(summary_sheet, 3, 3)
    for name, count in sorted(shift_counts.items(), key=lambda item: (-item[1], item[0])):
        employee = next((item for item in context.employees if item.name == name), None)
        summary_sheet.append([name, count, employee.role if employee else "?"])
    style_body(summary_sheet, 4, summary_sheet.max_row, 3)

    row_index = summary_sheet.max_row + 2
    summary_sheet.cell(row=row_index, column=1, value="PM Staff Days Off").font = subheader_font
    row_index += 1
    summary_sheet.append(["Employee", "Days Working", "Days Off"])
    style_header(summary_sheet, row_index, 3)
    pm_names = {employee.name for employee in context.employees if employee.role in ("leadership", "pm_staff")}
    by_employee = {}
    for assignment in assignments:
        if assignment.employee in pm_names and assignment.shift == "PM":
            by_employee.setdefault(assignment.employee, set()).add(assignment.day)
    for name in sorted(pm_names):
        working = by_employee.get(name, set())
        off = set(context.open_days) - working
        summary_sheet.append(
            [
                name,
                ", ".join(sorted(working, key=context.open_days.index)),
                ", ".join(sorted(off, key=context.open_days.index)) if off else "-",
            ]
        )
    style_body(summary_sheet, row_index + 1, summary_sheet.max_row, 3)

    row_index = summary_sheet.max_row + 2
    summary_sheet.cell(row=row_index, column=1, value=f"Validation: {report['status']}").font = Font(
        name="Arial",
        bold=True,
        size=12,
        color="C00000" if report["errors"] else "00B050",
    )
    summary_sheet.cell(row=row_index + 1, column=1, value=f"Errors: {len(report['errors'])}").font = body_font
    summary_sheet.cell(row=row_index + 2, column=1, value=f"Warnings: {len(report['warnings'])}").font = body_font
    for column_index, width in enumerate([18, 36, 36, 20], 1):
        summary_sheet.column_dimensions[get_column_letter(column_index)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
